from datetime import datetime
import pandas as pd
import numpy as np
from dash import dcc, html, Input, Output, State, dash_table
from django_plotly_dash import DjangoDash
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from scipy.stats import linregress
from plotly_integration.models import CIEFReport, CIEFMetadata, CIEFTimeSeries
from scipy.signal import find_peaks, savgol_filter, argrelextrema
import dash_bootstrap_components as dbc
from openpyxl import load_workbook
import io

app = DjangoDash("cIEFReportViewerApp")

app.layout = html.Div([
    dcc.Store(id="selected-result-ids"),
    dcc.Store(id="reduced-result-ids"),
    dcc.Store(id="nonreduced-result-ids"),
    dcc.Store(id="standard-regression-params"),
    dcc.Store(id="selected-report"),

    dcc.Tabs(id="main-tabs", value="tab-select-report", persistence=False, children=[

        # Select Report Tab
        dcc.Tab(label="Select Report", value="tab-select-report", children=[
            html.Div([
                dash_table.DataTable(
                    id="cief-report-table",
                    columns=[
                        {"name": "Report Name", "id": "report_name"},
                        {"name": "Project ID", "id": "project_id"},
                        {"name": "User ID", "id": "user_id"},
                        {"name": "Date Created", "id": "date_created"},
                        {"name": "Samples", "id": "num_samples"}
                    ],
                    data=[],
                    row_selectable="single",
                    filter_action="native",
                    sort_action="native",
                    page_size=20,
                    style_table={"height": "70vh", "overflowY": "auto"},
                    style_cell={"textAlign": "center", "padding": "8px"},
                    style_header={
                        "backgroundColor": "#e9f1fb",
                        "fontWeight": "bold",
                        "color": "#0047b3",
                        "borderBottom": "2px solid #0047b3"
                    }
                )
            ])
        ]),

        # Reduced Tab
        dcc.Tab(label="Electropherogram", value="electropherogram", children=[

            html.Div(style={"display": "flex"}, children=[
                html.Div([
                    dcc.Graph(
                        id="electropherogram",
                        config={'responsive': True},
                        style={"height": "100%"}
                    )
                ], style={"width": "80%", "padding": "10px", "minHeight": "1000px"}),

                html.Div([
                    html.H4("Marker Settings"),
                    html.Label("Marker RT (min):"),
                    dcc.Input(id="marker-rt", type="number", value=7, step=0.1, style={"width": "100%"}),

                    html.Label("Marker Label:"),
                    dcc.Input(id="marker-label", type="text", value="10 kDa",
                              style={"width": "100%", "marginBottom": "20px"}),

                    html.Hr(),
                    html.H4("Plot Settings"),
                    html.Label("Y- Axis Scaling"),
                    dcc.Input(id="reduced-y-axis-scaling", type="number", value=1.0, step=0.01,
                              style={"width": "100%"}),
                    html.Label("Subplot Vertical Spacing"),
                    dcc.Input(id="reduced-subplot-vertical-spacing", type="number", value=0.025, step=0.005,
                              style={"width": "100%"}),

                    html.Hr(),
                    html.H4("Peak Detection"),

                    html.Label("Skip Time After Marker (min):"),
                    dcc.Input(id="skip-time", type="number", value=0.3, step=0.05, style={"width": "100%"}),

                    html.Label("Max Peaks:"),
                    dcc.Input(id="max-peaks", type="number", value=4, step=1, min=1,
                              style={"width": "100%"}),

                    html.Label("Prominence Threshold:"),
                    dcc.Input(id="prominence-threshold", type="number", value=1.0, step=0.01,
                              style={"width": "100%"}),

                    html.Label("Valley Search Window (min):"),
                    dcc.Input(id="valley-search-window", type="number", value=1, step=0.1,
                              style={"width": "100%"}),

                    html.Label("Valley Drop Ratio (0–1):"),
                    dcc.Input(id="valley-drop-ratio", type="number", value=0.2, step=0.05, min=0, max=1,
                              style={"width": "100%"}),

                    html.Label("Smoothing Window (odd integer):"),
                    dcc.Input(id="smoothing-window", type="number", value=11, step=2, min=3,
                              style={"width": "100%"}),

                    html.Label("Smoothing Polyorder:"),
                    dcc.Input(id="smoothing-polyorder", type="number", value=3, step=1, min=1,
                              style={"width": "100%", "marginBottom": "20px"}),

                    html.Hr(),
                    html.H4("Light Chain Timing"),
                    html.Button("Calculate Light Chain Time", id="calc-light-chain-btn",
                                style={"width": "100%", "marginBottom": "10px"}),
                    dcc.Input(id="light-chain-time", type="number", placeholder="Light Chain Time (min)",
                              readOnly=True, style={"width": "100%"})

                ], style={"width": "20%", "padding": "10px"})
            ])
        ]),
        dcc.Tab(label="Results Table", value="tab-table", children=[
            html.Div([
                html.H4("cIEF Results Table", style={'textAlign': 'center', 'color': '#0056b3'}),
                dash_table.DataTable(
                    id="cief-table",
                    columns=[],  # will be filled by callback
                    data=[],
                    style_header={
                        'backgroundColor': '#0056b3',  # blue
                        'color': 'white',
                        'fontWeight': 'bold',
                        'textAlign': 'center'
                    },
                    style_table={"overflowX": "auto"},
                    style_cell={"textAlign": "center"},
                ),
                html.Button("Export cIEF Table", id="export-cief-btn"),
                dcc.Download(id="download-cief-xlsx")
            ]),

        ]),
    ])
])


@app.callback(
    Output("cief-report-table", "data"),
    Input("main-tabs", "value")
)
def load_report_table(tab_value):
    if tab_value != "tab-select-report":
        return []
    reports = CIEFReport.objects.order_by("-date_created")
    return [{
        "report_name": r.report_name,
        "project_id": r.project_id,
        "user_id": r.user_id,
        "date_created": r.date_created.strftime("%Y-%m-%d %H:%M"),
        "num_samples": len(r.selected_result_ids.split(","))
    } for r in reports]


@app.callback(
    Output("selected-result-ids", "data"),
    Output("selected-report", "data"),
    Input("cief-report-table", "selected_rows"),
    State("cief-report-table", "data")
)
def store_selected_result_ids(selected_rows, table_data):
    if selected_rows:
        row = table_data[selected_rows[0]]
        report = CIEFReport.objects.filter(report_name=row["report_name"]).first()
        report_name = report.report_name
        if report:
            return [r.strip() for r in report.selected_result_ids.split(",")], report_name
    return [], []


def detect_valley_to_valley_peaks_with_standards(
        df,
        signal_col="channel_1",
        time_col="time_min",
        baseline_cutoff_time=1.0,
        prominence_threshold=1.0,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3,
):
    from scipy.signal import savgol_filter, find_peaks

    time = df[time_col].values
    signal = df[signal_col].values
    smoothed = savgol_filter(signal, smoothing_window, smoothing_polyorder)

    interval = time[1] - time[0]
    min_distance = int(0.3 / interval)
    peak_indices, _ = find_peaks(smoothed, prominence=prominence_threshold, distance=min_distance)

    peak_infos = []
    for idx in peak_indices:
        peak_time = time[idx]
        peak_height = smoothed[idx]
        min_valley_height = peak_height * (1 - valley_drop_ratio)

        left_limit = max(0, idx - int(valley_search_window / interval))
        left_valley_idx = left_limit + np.argmin(smoothed[left_limit:idx])
        left_valley_val = smoothed[left_valley_idx]

        right_limit = min(len(smoothed) - 1, idx + int(valley_search_window / interval))
        right_valley_idx = idx + np.argmin(smoothed[idx:right_limit + 1])
        right_valley_val = smoothed[right_valley_idx]

        if left_valley_val > min_valley_height or right_valley_val > min_valley_height:
            continue

        baseline = np.linspace(left_valley_val, right_valley_val, right_valley_idx - left_valley_idx + 1)
        signal_segment = smoothed[left_valley_idx:right_valley_idx + 1]
        time_segment = time[left_valley_idx:right_valley_idx + 1]
        area = np.trapz(signal_segment - baseline, time_segment)

        if area > 0:
            peak_infos.append({
                "peak_time": time[idx],
                "peak_height": peak_height,
                "area": area,
                "start_time": time[left_valley_idx],
                "end_time": time[right_valley_idx],
                "baseline": baseline,
                "baseline_time": time_segment,
                "peak_index": idx
            })

    peak_infos = sorted(peak_infos, key=lambda x: x["peak_time"])
    std_candidates = [p for p in peak_infos if p["peak_time"] > baseline_cutoff_time]

    if len(std_candidates) < 4:
        return peak_infos, None, None, []

    first_two = std_candidates[:2]
    last_two = std_candidates[-2:]
    standards = first_two + last_two

    pi_values = [10.0, 9.5, 5.5, 4.0]
    for peak, pi in zip(standards, pi_values):
        peak["pI"] = pi

    rt_vals = np.array([p["peak_time"] for p in standards])
    pi_vals = np.array([p["pI"] for p in standards])
    slope, intercept, *_ = linregress(rt_vals, pi_vals)

    return peak_infos, slope, intercept, standards


def shade_peak(fig, df, start_time, end_time, baseline, row, col, label, peak_time, peak_height, class_label):
    PEAK_CLASS_COLORS = {
        "LMW": "rgba(76, 175, 80, 0.3)",  # green (complement of red)
        "Intact": "rgba(33, 150, 243, 0.3)",  # blue (cool, clinical)
        "Light Chain": "rgba(33, 150, 243, 0.3)",  # blue (cool, clinical)
        "Heavy Chain": "rgba(255, 193, 7, 0.3)",  # amber (warm, eye-catching)
        "HMW": "rgba(244, 67, 54, 0.3)",  # red (danger, HMW)

    }

    color = PEAK_CLASS_COLORS.get(class_label, "rgba(0,100,200,0.2)")  # fallback gray

    region = df[(df["time_min"] >= start_time) & (df["time_min"] <= end_time)].copy()
    if region.empty:
        return

    fig.add_trace(
        go.Scatter(
            x=np.concatenate([region["time_min"], region["time_min"][::-1]]),
            y=np.concatenate([region["channel_1"], [baseline] * len(region)]),
            fill="toself",
            fillcolor=color,
            line=dict(color="rgba(255,255,255,0)"),
            showlegend=False,
            hoverinfo="skip"
        ),
        row=row,
        col=col
    )

    fig.add_annotation(
        x=peak_time,
        y=peak_height,
        text=label,
        showarrow=True,
        arrowhead=1,
        ax=0,
        ay=-30,
        row=row,
        col=col
    )


def generate_chromatogram_figure_advanced(
        result_df_by_id,
        title="Chromatograms",
        marker_rt=None,
        marker_label="10 kDa",
        skip_time=0.3,
        max_peaks=4,
        prominence_threshold=0.05,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3,
        light_chain_time=None,
        table_output=None,
        regression_slope=None,
        regression_intercept=None,
        y_scale=1,
        subplot_vertical_spacing=0.25,
        baseline_cutoff_time=10.0,
):
    num_rows = len(result_df_by_id)

    fig = make_subplots(
        rows=num_rows, cols=1,
        shared_xaxes=False,
        vertical_spacing=subplot_vertical_spacing,
        subplot_titles=[meta["sample_id"] for meta in result_df_by_id.values()]
    )

    for i, (mid, meta) in enumerate(result_df_by_id.items(), start=1):
        df = meta["data"]
        if df.empty:
            continue

        fig.add_trace(
            go.Scatter(x=df["time_min"], y=df["channel_1"], mode="lines", name=meta["sample_id"]),
            row=i, col=1
        )

        # Marker annotation
        marker_peak_time = None
        if marker_rt is not None:
            df_marker = df[(df["time_min"] >= marker_rt - 0.5) & (df["time_min"] <= marker_rt + 0.5)]
            if not df_marker.empty:
                idx = df_marker["channel_1"].idxmax()
                marker_peak_time = df_marker.loc[idx, "time_min"]
                peak_height = df_marker.loc[idx, "channel_1"]

                fig.add_annotation(
                    x=marker_peak_time,
                    y=peak_height,
                    text=f"{marker_label} ({marker_peak_time:.2f} min)",
                    showarrow=True,
                    arrowhead=1,
                    ax=0,
                    ay=-40,
                    row=i, col=1
                )

        # Peak detection
        df_after_marker = df[df["time_min"] > (marker_peak_time or 0) + skip_time].copy()
        # Detect peaks & standard pI regression
        peaks, slope, intercept, standards = detect_valley_to_valley_peaks_with_standards(
            df,
            baseline_cutoff_time=baseline_cutoff_time,
            prominence_threshold=prominence_threshold,
            valley_search_window=valley_search_window,
            valley_drop_ratio=valley_drop_ratio,
            smoothing_window=smoothing_window,
            smoothing_polyorder=smoothing_polyorder,
        )

        print(f'peaks: {peaks}')
        print(f'standards: {standards}')
        print(f'slope: {slope}')
        print(f'intercept: {intercept}')

        total_area = sum(p["area"] for p in peaks)
        peaks = sorted(peaks, key=lambda x: x["peak_height"], reverse=True)

        light_chain = heavy_chain = None
        if len(peaks) >= 2:
            top2 = peaks[:2]
            if top2[0]["peak_time"] < top2[1]["peak_time"]:
                light_chain, heavy_chain = top2[0], top2[1]
            else:
                light_chain, heavy_chain = top2[1], top2[0]

        percentages = {"LMW": 0.0, "Light Chain": 0.0, "Heavy Chain": 0.0, "HMW": 0.0}

        for p in peaks:
            # Y Axes Scaling
            if not y_scale or y_scale == 1:
                autoscale = True
                fig.update_yaxes(title_text="UV", row=i, col=1)
            else:
                autoscale = False
                max_signal = df["channel_1"].max()
                y_max = max_signal * y_scale
                fig.update_yaxes(
                    title_text="UV",
                    autorange=True,
                    autorangeoptions=dict(maxallowed=y_max),
                    row=i,
                    col=1
                )
            fig.update_xaxes(title_text="Time (min)", row=i, col=1)

            # Peak Classification and percentage calculation
            pct = (p["area"] / total_area * 100) if total_area else 0
            label = f"{pct:.1f}%"
            class_label = None

            if p == light_chain:
                label = f"Light Chain<br>({pct:.1f}%)"
                class_label = "Light Chain"
            elif p == heavy_chain:
                label = f"Heavy Chain<br>({pct:.1f}%)"
                class_label = "Heavy Chain"
            elif light_chain and p["peak_time"] < light_chain["peak_time"]:
                label = f"LMW<br>({pct:.1f}%)"
                class_label = "LMW"
            elif heavy_chain and p["peak_time"] > heavy_chain["peak_time"]:
                label = f"HMW<br>({pct:.1f}%)"
                class_label = "HMW"
            else:
                continue

            # Optional: Calculate MW
            if regression_slope is not None and regression_intercept is not None:
                log_mw = regression_slope * p["peak_time"] + regression_intercept
                mw_kda = np.exp(log_mw)
                label += f"<br>{mw_kda:.1f} kDa"

            percentages[class_label] += pct

            if autoscale:
                peak_height = p["peak_height"]
            elif autoscale == False and p["peak_height"] < y_max:
                peak_height = p["peak_height"]
            else:
                peak_height = y_max

            shade_peak(
                fig, df,
                start_time=p["start_time"],
                end_time=p["end_time"],
                baseline=p["baseline"][0],
                row=i, col=1,
                label=label,
                peak_time=p["peak_time"],
                peak_height=peak_height,
                class_label=class_label
            )

        # Append summary to table output
        if table_output is not None:
            table_output.append({
                "Sample Name": meta["sample_id"],
                "LMW (%)": round(percentages["LMW"], 1),
                "Light Chain (%)": round(percentages["Light Chain"], 1),
                "Heavy Chain (%)": round(percentages["Heavy Chain"], 1),
                "HMW (%)": round(percentages["HMW"], 1),
            })

    fig.update_layout(
        height=300 * num_rows,
        title=title,
        showlegend=False,
        template="plotly_white",
        margin=dict(t=40, b=40, l=40, r=30)
    )
    print(f'table output: {table_output}')
    return fig, table_output




@app.callback(
    [Output("electropherogram", "figure"),
     Output("electropherogram", "config"),
     Output("cief-table", "data"),
     Output("cief-table", "columns")],

    [
        Input("selected-result-ids", "data"),
        Input("marker-rt", "value"),
        Input("marker-label", "value"),
        Input("skip-time", "value"),
        Input("max-peaks", "value"),
        Input("prominence-threshold", "value"),
        Input("valley-search-window", "value"),
        Input("valley-drop-ratio", "value"),
        Input("smoothing-window", "value"),
        Input("smoothing-polyorder", "value"),
        Input("light-chain-time", "value"),
        Input("standard-regression-params", "data"),
        Input("reduced-y-axis-scaling", "value"),
        Input("reduced-subplot-vertical-spacing", "value"),
        State("selected-report", "data"),

    ]
)
def reduced_callback(result_ids, marker_rt, marker_label, skip_time, max_peaks,
                     prominence_threshold, valley_search_window, valley_drop_ratio,
                     smoothing_window, smoothing_polyorder, light_chain_time,
                     regression_params, y_scale, subplot_vertical_spacing, selected_report):
    metas = CIEFMetadata.objects.filter(id__in=result_ids)
    result_df_by_id = {
        str(m.id): {
            "sample_id": m.sample_id_full,
            "data": pd.DataFrame(list(
                CIEFTimeSeries.objects.filter(metadata_id=m.id).values("time_min", "channel_1")
            )).sort_values("time_min")
        }
        for m in metas
    }

    # ✅ Initialize table_output as empty list
    table_output = []

    slope = regression_params.get("slope") if regression_params else None
    intercept = regression_params.get("intercept") if regression_params else None

    fig, table_data = generate_chromatogram_figure_advanced(
        result_df_by_id,
        title="Reduced Chromatograms with Peak Classification",
        marker_rt=marker_rt,
        marker_label=marker_label,
        skip_time=skip_time,
        max_peaks=max_peaks,
        prominence_threshold=prominence_threshold,
        valley_search_window=valley_search_window,
        valley_drop_ratio=valley_drop_ratio,
        smoothing_window=smoothing_window,
        smoothing_polyorder=smoothing_polyorder,
        light_chain_time=light_chain_time,
        table_output=table_output,
        regression_slope=slope,
        regression_intercept=intercept,
        y_scale=y_scale,
        subplot_vertical_spacing=subplot_vertical_spacing,
        baseline_cutoff_time=10.0,
    )
    print(f'table output: {table_output}')

    columns = [{"name": k, "id": k} for k in table_data[0].keys()] if table_data else []

    plot_config = {
        'toImageButtonOptions': {
            'filename': f"{datetime.now().strftime('%Y%m%d')}-R-{selected_report}",
            'format': 'png',
            # 'height': 600,
            # 'width': 800,
            # 'scale': 2
        }}
    return fig, plot_config, table_data, columns


@app.callback(
    Output("download-cief-xlsx", "data"),
    Input("export-cief-btn", "n_clicks"),
    State("reduced-table", "data"),
    State("selected-report", "data"),
    prevent_initial_call=True
)
def export_table(n_clicks, table_data, report_name):
    filename = f"{report_name}_Reduced.xlsx"
    df = pd.DataFrame(table_data)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1)
        worksheet = writer.sheets["Sheet1"]
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        worksheet.cell(row=1, column=1).value = "Reduced Results"

    buffer.seek(0)

    def write_buffer(out_io):
        out_io.write(buffer.getvalue())

    return dcc.send_bytes(write_buffer, filename)
