from datetime import datetime
from math import ceil

import pandas as pd
import numpy as np
from dash import dcc, html, Input, Output, State, dash_table
from django_plotly_dash import DjangoDash
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from scipy.stats import linregress
from plotly_integration.models import CIEFReport, CIEFMetadata, CIEFTimeSeries
from scipy.signal import find_peaks, savgol_filter, argrelextrema
import dash_bootstrap_components as dbc
from openpyxl import load_workbook
import io
import plotly.colors as pc
import dash_bootstrap_components as dbc

app = DjangoDash("cIEFReportViewerApp")

app.layout = html.Div([
    dcc.Store(id="selected-result-ids"),
    dcc.Store(id="reduced-result-ids"),
    dcc.Store(id="nonreduced-result-ids"),
    dcc.Store(id="standard-regression-params"),
    dcc.Store(id="selected-report"),

    dcc.Tabs(id="main-tabs", value="tab-select-report", persistence=False, children=[

        # Select Report Tab
        dcc.Tab(label="Select Report", value="tab-select-report", children=[
            html.Div([
                dash_table.DataTable(
                    id="cief-report-table",
                    columns=[
                        {"name": "Report Name", "id": "report_name"},
                        {"name": "Project ID", "id": "project_id"},
                        {"name": "User ID", "id": "user_id"},
                        {"name": "Date Created", "id": "date_created"},
                        {"name": "Samples", "id": "num_samples"}
                    ],
                    data=[],
                    row_selectable="single",
                    filter_action="native",
                    sort_action="native",
                    page_size=20,
                    style_table={"height": "70vh", "overflowY": "auto"},
                    style_cell={"textAlign": "center", "padding": "8px"},
                    style_header={
                        "backgroundColor": "#e9f1fb",
                        "fontWeight": "bold",
                        "color": "#0047b3",
                        "borderBottom": "2px solid #0047b3"
                    }
                )
            ])
        ]),

        # Plots
        dcc.Tab(label="Electropherogram", value="electropherogram", children=[

            html.Div(style={"display": "flex"}, children=[
                html.Div([
                    dcc.Graph(
                        id="electropherogram",
                        config={'responsive': True},
                        style={"height": "100%"}
                    )
                ], style={"width": "80%", "padding": "10px", "minHeight": "1000px"}),

                html.Div([
                    html.H4("STD Detection Settings"),
                    html.Label("STD Start Cutoff Time (min):"),
                    dcc.Input(id="std-cutoff-time", type="number", value=12, step=0.1, style={"width": "100%"}),

                    html.Label("STD End Cutoff Time (min):"),
                    dcc.Input(id="std-end-cutoff-time", type="number", value=30, step=0.1,
                              style={"width": "100%"}),

                    html.Label("Prominence Threshold:"),
                    dcc.Input(id="std-promincence-threshold", type="number", value=7000, step=0.1,
                              style={"width": "100%"}),

                    html.Label("Valley Search Window (min):"),
                    dcc.Input(id="std-valley-search-window", type="number", value=1, step=0.01,
                              style={"width": "100%"}),

                    html.Div([
                        html.Label("Show STD Regression Curve:"),
                        dbc.Checkbox(
                            id="show-std-regression",
                            value=False,  # default unchecked
                        )
                    ]),

                    html.Hr(),
                    html.H4("Plot Settings"),
                    html.Div([
                        html.Label("Shade Peaks:"),
                        dbc.Checkbox(
                            id="shade-peaks",
                            value=False,  # default checked
                        )
                    ]),
                    html.Label("AutoScale X Axis:"),
                    dcc.Checklist(
                        options=[{"label": "", "value": True}],
                        value=[True],
                        id="x-axis-autoscale",
                    ),
                    html.Label("Number of Columns:"),
                    dcc.Input(
                        id="num-subplot-cols",
                        type="number",
                        value=3,  # default
                        min=1,
                        step=1,
                        style={"width": "100%"}
                    ),
                    html.Label("X Axis Min:"),
                    dcc.Input(id="x-axis-min", type="number", value=0, step=0.01,
                              style={"width": "100%"}),
                    html.Label("X Axis Max:"),
                    dcc.Input(id="x-axis-max", type="number", value=40, step=0.01,
                              style={"width": "100%"}),
                    html.Label("Y- Axis Scaling"),
                    dcc.Input(id="y-axis-scaling", type="number", value=1.0, step=0.01,
                              style={"width": "100%"}),
                    html.Label("Subplot Vertical Spacing"),
                    dcc.Input(id="subplot-vertical-spacing", type="number", value=0.05, step=0.005,
                              style={"width": "100%"}),
                    html.Label("Subplot Horizontal Spacing"),
                    dcc.Input(
                        id="subplot-horizontal-spacing",
                        type="number",
                        value=0.05,  # default value
                        step=0.01,
                        min=0.01,
                        style={"width": "100%"}
                    ),

                    html.Hr(),
                    html.H4("Peak Detection"),

                    html.Label("pI Calculation Method:"),
                    dcc.Dropdown(
                        id="pi-method",
                        options=[
                            {"label": "Peak Max RT", "value": "peak_max"},
                            {"label": "Weighted RT", "value": "weighted_rt"}
                        ],
                        value="peak_max",
                        style={"width": "100%", "marginBottom": "20px"}
                    ),

                    html.Label("Max Peaks:"),
                    dcc.Input(id="max-peaks", type="number", value=3, step=1, min=1,
                              style={"width": "100%"}),

                    html.Label("Prominence Threshold:"),
                    dcc.Input(id="prominence-threshold", type="number", value=1000, step=0.01,
                              style={"width": "100%"}),

                    html.Label("Valley Search Window (min):"),
                    dcc.Input(id="valley-search-window", type="number", value=0.7, step=0.01,
                              style={"width": "100%"}),

                    html.Label("Valley Drop Ratio (0–1):"),
                    dcc.Input(id="valley-drop-ratio", type="number", value=0.3, step=0.05, min=0, max=1,
                              style={"width": "100%"}),

                    html.Label("Smoothing Window (odd integer):"),
                    dcc.Input(id="smoothing-window", type="number", value=3, step=2, min=3,
                              style={"width": "100%"}),

                    html.Label("Smoothing Polyorder:"),
                    dcc.Input(id="smoothing-polyorder", type="number", value=1, step=1, min=1,
                              style={"width": "100%", "marginBottom": "20px"}),

                ], style={"width": "20%", "padding": "10px"})
            ])
        ]),
        dcc.Tab(label="Results Table", value="tab-table", children=[
            html.Div([
                html.H4("cIEF Results Table", style={'textAlign': 'center', 'color': '#0056b3'}),
                dash_table.DataTable(
                    id="cief-table",
                    columns=[],  # will be filled by callback
                    data=[],
                    style_header={
                        'backgroundColor': '#0056b3',  # blue
                        'color': 'white',
                        'fontWeight': 'bold',
                        'textAlign': 'center'
                    },
                    style_table={"overflowX": "auto"},
                    style_cell={"textAlign": "center"},
                ),
                html.Button("Export cIEF Table", id="export-cief-btn"),
                dcc.Download(id="download-cief-xlsx")
            ]),

        ]),
    ])
])


@app.callback(
    Output("cief-report-table", "data"),
    Input("main-tabs", "value")
)
def load_report_table(tab_value):
    if tab_value != "tab-select-report":
        return []
    reports = CIEFReport.objects.order_by("-date_created")
    return [{
        "report_name": r.report_name,
        "project_id": r.project_id,
        "user_id": r.user_id,
        "date_created": r.date_created.strftime("%Y-%m-%d %H:%M"),
        "num_samples": len(r.selected_result_ids.split(","))
    } for r in reports]


@app.callback(
    Output("selected-result-ids", "data"),
    Output("selected-report", "data"),
    Input("cief-report-table", "selected_rows"),
    State("cief-report-table", "data")
)
def store_selected_result_ids(selected_rows, table_data):
    if selected_rows:
        row = table_data[selected_rows[0]]
        report = CIEFReport.objects.filter(report_name=row["report_name"]).first()
        report_name = report.report_name
        if report:
            return [r.strip() for r in report.selected_result_ids.split(",")], report_name
    return [], []


def detect_valley_to_valley_peaks(
        df,
        signal_col="channel_1",
        time_col="time_min",
        max_peaks=3,
        prominence_threshold=25.0,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3,

):
    """
    Detects peaks using valley-to-valley integration with adaptive valley thresholding.
    Returns a list of dictionaries with peak info.
    """
    from scipy.signal import savgol_filter, find_peaks

    time = df[time_col].values
    signal = df[signal_col].values

    n_points = len(signal)
    safe_window = min(smoothing_window, n_points - 1 if n_points % 2 == 0 else n_points)
    if safe_window < 3:
        return [], []
    if safe_window % 2 == 0:
        safe_window -= 1

    smoothed = savgol_filter(signal, window_length=safe_window, polyorder=smoothing_polyorder)

    interval = time[1] - time[0]
    min_distance = max(1, int(0.3 / interval))
    peak_indices, _ = find_peaks(smoothed, prominence=prominence_threshold, distance=min_distance)

    peak_indices = sorted(peak_indices, key=lambda i: smoothed[i], reverse=True)[:max_peaks]
    peak_indices.sort()

    peak_infos = []

    for idx in peak_indices:
        peak_time = time[idx]
        peak_height = smoothed[idx]
        min_valley_height = peak_height * (1 - valley_drop_ratio)

        left_limit = max(0, idx - int(valley_search_window / interval))
        left_slice = smoothed[left_limit:idx]
        if left_slice.size == 0:
            continue
        left_valley_idx = left_limit + np.argmin(left_slice)
        left_valley_val = smoothed[left_valley_idx]

        right_limit = min(len(smoothed) - 1, idx + int(valley_search_window / interval))
        right_slice = smoothed[idx:right_limit + 1]
        if right_slice.size == 0:
            continue
        right_valley_idx = idx + np.argmin(right_slice)
        right_valley_val = smoothed[right_valley_idx]

        if left_valley_val > min_valley_height or right_valley_val > min_valley_height:
            continue

        baseline = np.linspace(left_valley_val, right_valley_val, right_valley_idx - left_valley_idx + 1)
        signal_segment = smoothed[left_valley_idx:right_valley_idx + 1]
        time_segment = time[left_valley_idx:right_valley_idx + 1]
        area = np.trapz(signal_segment - baseline, time_segment)

        if area > 0:
            peak_infos.append({
                "peak_time": time[idx],
                "peak_height": peak_height,
                "area": area,
                "start_time": time[left_valley_idx],
                "end_time": time[right_valley_idx],
                "baseline": baseline,
                "baseline_time": time_segment,
                "signal_segment": signal_segment,
                "peak_index": idx
            })

    return sorted(peak_infos, key=lambda x: x["area"], reverse=True), smoothed


def get_color_from_rt(rt, rt_min, rt_max, colorscale="Viridis"):
    norm = (rt - rt_min) / (rt_max - rt_min) if rt_max > rt_min else 0.5
    scale = pc.get_colorscale(colorscale)
    return pc.sample_colorscale(scale, norm)[0]


def generate_chromatogram_figure_advanced(
        result_df_by_id,
        title=None,
        max_peaks=4,
        prominence_threshold=0.05,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3,
        table_output=None,
        y_scale=1,
        subplot_vertical_spacing=0.25,
        subplot_horizontal_spacing=0.025,
        baseline_cutoff_time=10,
        x_axis_autoscale=True,
        x_axis_min=None,
        x_axis_max=None,
        plot_std_regression=True,
        std_prominence_threshold=10000,
        std_valley_search_window=1.0,
        std_end_cutoff_time=30,
        pi_method=None,
        shade_peaks=True,
        num_subplot_cols=1

):
    if len(result_df_by_id) == 0:
        return go.Figure(), []
    num_plots = len(result_df_by_id)
    num_cols = max(1, min(num_subplot_cols, num_plots))
    num_rows = ceil(num_plots / num_cols)

    titles = [meta["sample_id"] for meta in result_df_by_id.values()]
    specs = [[{"secondary_y": True} for _ in range(num_cols)] for _ in range(num_rows)]

    fig = make_subplots(
        rows=num_rows,
        cols=num_cols,
        shared_xaxes=False,
        vertical_spacing=subplot_vertical_spacing,
        horizontal_spacing=subplot_horizontal_spacing,
        subplot_titles=titles,
        specs=specs
    )

    for i, (mid, meta) in enumerate(result_df_by_id.items(), start=1):
        row = (i - 1) // num_cols + 1
        col = (i - 1) % num_cols + 1

        df = meta["data"].sort_values("time_min")
        fig.add_trace(go.Scatter(x=df["time_min"], y=df["channel_1"], mode="lines", name=meta["sample_id"]),
                      row=row, col=col)

        # STD Peak Detection
        std_number_front_peaks = 2
        std_number_back_peaks = 2

        # Detect all forward peaks
        forward_std_peaks, _ = detect_valley_to_valley_peaks(
            df[df["time_min"] > baseline_cutoff_time],
            signal_col="channel_1",
            time_col="time_min",
            max_peaks=10,
            prominence_threshold=std_prominence_threshold,
            valley_search_window=std_valley_search_window,
            valley_drop_ratio=valley_drop_ratio,
            smoothing_window=smoothing_window,
            smoothing_polyorder=smoothing_polyorder
        )

        standards_front = sorted(forward_std_peaks, key=lambda p: p["peak_time"])[:2]

        # Reverse DataFrame for backward peak search
        df_rev = df[df["time_min"] > std_end_cutoff_time]
        df_rev = df_rev.iloc[::-1].copy()
        t_min, t_max = df_rev["time_min"].min(), df_rev["time_min"].max()
        df_rev["time_min"] = t_max - (df_rev["time_min"] - t_min)  # flip time axis

        backward_std_peaks, _ = detect_valley_to_valley_peaks(
            df_rev,
            signal_col="channel_1",
            time_col="time_min",
            max_peaks=10,
            prominence_threshold=std_prominence_threshold,
            valley_search_window=std_valley_search_window,
            valley_drop_ratio=valley_drop_ratio,
            smoothing_window=smoothing_window,
            smoothing_polyorder=smoothing_polyorder
        )

        # Flip times back to original
        for p in backward_std_peaks:
            p["peak_time"] = t_max - (p["peak_time"] - t_min)
            p["start_time"] = t_max - (p["start_time"] - t_min)
            p["end_time"] = t_max - (p["end_time"] - t_min)
            p["baseline_time"] = t_max - (p["baseline_time"] - t_min)

        standards_back = sorted(backward_std_peaks, key=lambda p: p["peak_time"], reverse=True)[:2]

        # Combine & assign pI values
        standard_peaks = sorted(standards_front + standards_back, key=lambda p: p["peak_time"])
        pi_values = [10.0, 9.5, 5.5, 4.0]
        for peak, pi in zip(standard_peaks, pi_values):
            peak["pI"] = pi

        if len(standard_peaks) == std_number_front_peaks + std_number_back_peaks:
            sample_start = standard_peaks[1]["end_time"] + 0.25  # end of 2nd standard
            sample_end = standard_peaks[2]["start_time"] - 1  # start of 3rd standard
            sample_df = df[(df["time_min"] >= sample_start) & (df["time_min"] <= sample_end)]
            print(f"Sample region from {sample_start:.2f} to {sample_end:.2f} min")
        elif 1 < len(standards_back) < std_number_back_peaks:
            sample_start = standard_peaks[1]["end_time"] + 0.25  # end of 2nd standard
            sample_end = standard_peaks[-1]["start_time"] - 1  # start of 3rd standard
            sample_df = df[(df["time_min"] >= sample_start) & (df["time_min"] <= sample_end)]
            print(f"Sample region from {sample_start:.2f} to {sample_end:.2f} min. Error with Back Standard Peaks")

        elif len(standards_back) == 0:
            sample_start = standard_peaks[1]["end_time"] + 0.25  # end of 2nd standard
            sample_end = df["time_min"].max()
            sample_df = df[(df["time_min"] >= sample_start) & (df["time_min"] <= sample_end)]
            print(f"Sample region from {sample_start:.2f} to {sample_end:.2f} min. Error with Back Standard Peaks")
        elif len(standards_front) < std_number_front_peaks:
            sample_start = standard_peaks[0]["end_time"] + 0.25  # end of 2nd standard
            sample_end = standard_peaks[-1]["start_time"] - 1  # start of 3rd standard
            sample_df = df[(df["time_min"] >= sample_start) & (df["time_min"] <= sample_end)]
            print(f"Sample region from {sample_start:.2f} to {sample_end:.2f} min. Error with Back Standard Peaks")
        else:
            sample_df = df.iloc[0:0]

        for std in standard_peaks:
            fig.add_annotation(
                x=std["peak_time"],
                y=std["peak_height"],
                text=f"pI {std['pI']}",
                showarrow=True,
                arrowhead=2,
                ax=0,
                ay=-40,
                row=row,
                col=col
            )

        if len(standard_peaks) >= 2:
            rt_vals = np.array([p["peak_time"] for p in standard_peaks])
            pi_vals = np.array([p["pI"] for p in standard_peaks])

            # Linear regression: pI = slope * RT + intercept
            slope, intercept, r_val, *_ = linregress(rt_vals, pi_vals)

            x_range = np.linspace(df["time_min"].min(), df["time_min"].max(), 200)
            y_fit = slope * x_range + intercept

            if plot_std_regression:
                fig.add_trace(
                    go.Scatter(
                        x=rt_vals,
                        y=pi_vals,
                        mode="markers+text",
                        name="Standard pI",
                        # text=[f"{pi:.1f}" for pi in pi_vals],
                        textposition="top center",
                        marker=dict(size=8, color="black", symbol="circle"),
                    ),
                    row=row, col=col, secondary_y=True  # 👈 this line is critical
                )

                fig.add_trace(
                    go.Scatter(
                        x=x_range,
                        y=y_fit,
                        mode="lines",
                        name="pI Regression",
                        line=dict(color="red", dash="dash"),
                    ),
                    row=row, col=col, secondary_y=True  # 👈 also critical
                )

                fig.update_yaxes(
                    title_text="pI",
                    range=[3, 12],
                    secondary_y=True,
                    row=row,
                    col=col,
                    tickfont=dict(color="red"),
                    title_font=dict(color="red")
                )

                # Setup secondary y-axis
                fig.update_layout({f'yaxis{i + 1}2': dict(
                    overlaying=f'y{i + 1}',
                    side='right',
                    title=dict(text="pI", font=dict(color="red")),
                    showgrid=False,
                    tickfont=dict(color="red")
                )})

        # Detect sample peaks
        sample_peaks, _ = detect_valley_to_valley_peaks(
            sample_df,
            signal_col="channel_1",
            time_col="time_min",
            max_peaks=max_peaks,
            prominence_threshold=prominence_threshold,
            valley_search_window=valley_search_window,
            valley_drop_ratio=valley_drop_ratio,
            smoothing_window=smoothing_window,
            smoothing_polyorder=smoothing_polyorder,
        )

        total_area = sum(p["area"] for p in sample_peaks)
        sample_peaks = sorted(sample_peaks, key=lambda x: x["peak_height"], reverse=True)

        if x_axis_autoscale and sample_peaks:
            peaks = sorted(sample_peaks, key=lambda p: p["peak_time"])
            x_start = peaks[0]["start_time"] - 0.5
            x_end = peaks[-1]["end_time"] + 1
            fig.update_xaxes(range=[x_start, x_end], row=row, col=col)
        elif x_axis_autoscale and not sample_peaks:
            x_start = standard_peaks[0]["start_time"] - 0.5
            x_end = standard_peaks[-1]["end_time"] + 1
            fig.update_xaxes(range=[x_start, x_end], row=row, col=col)
        else:
            fig.update_xaxes(range=[x_axis_min, x_axis_max], row=row, col=col)

        if sample_peaks:
            rt_min = min(p["peak_time"] for p in sample_peaks)
            rt_max = max(p["peak_time"] for p in sample_peaks)
            # rt_min = standard_peaks[1]["peak_time"]
            # rt_max = standard_peaks[-1]["peak_time"] - 10

            for p in sample_peaks:
                signal_adj = p["signal_segment"] - p["baseline"]
                weighted_rt = np.sum(p["baseline_time"] * signal_adj) / np.sum(signal_adj)

                rt_for_pi = {
                    "peak_max": p["peak_time"],
                    "weighted_rt": weighted_rt
                }.get(pi_method, p["peak_time"])

                pI_value = round(slope * rt_for_pi + intercept, 2)
                pct = (p["area"] / total_area * 100) if total_area else 0

                color = None
                if shade_peaks:
                    color = get_color_from_rt(p["peak_time"], rt_min, rt_max)

                region = df[(df["time_min"] >= p["start_time"]) & (df["time_min"] <= p["end_time"])]
                if not region.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=np.concatenate([region["time_min"], region["time_min"][::-1]]),
                            y=np.concatenate([region["channel_1"], [p["baseline"][0]] * len(region)]),
                            fill="toself",
                            fillcolor=color,
                            line=dict(color="rgba(255,255,255,0)"),
                            showlegend=False,
                            hoverinfo="skip"
                        ),
                        row=row, col=col
                    )

                fig.add_annotation(
                    x=p["peak_time"],
                    y=p["peak_height"],
                    text=f"{pI_value} pI<br>{pct:.1f}%",
                    showarrow=True,
                    arrowhead=1,
                    ax=0,
                    ay=-30,
                    row=row,
                    col=col
                )
                if table_output is not None:
                    table_output.append({
                        "Sample Name": meta["sample_id"],
                        "pI": pI_value,
                        "%Area": round(pct,2)
                    })

                    fig.update_layout(
                        height=300 * num_rows,
                        title=title,
                        showlegend=False,
                        template="plotly_white",
                        margin=dict(t=40, b=40, l=40, r=30)
                    )
                    print(f'table output: {table_output}')

    #Data Table Creation Logic
    # After processing all samples, convert the list to a DataFrame
    peak_df = pd.DataFrame(table_output)

    if peak_df.empty:
        return fig, []

    # Group by 'Sample Name' and aggregate as needed
    grouped_df = peak_df.groupby('Sample Name').agg({
        'pI': list,
        '%Area': list
    }).reset_index()

    grouped_df['pI'] = grouped_df['pI'].apply(lambda x: ', '.join(f"{pi:.2f}" for pi in x))
    grouped_df['%Area'] = grouped_df['%Area'].apply(lambda x: ', '.join(f"{area:.1f}%" for area in x))

    data = grouped_df.to_dict('records')

    return fig, data


@app.callback(
    [Output("electropherogram", "figure"),
     Output("electropherogram", "config"),
     Output("cief-table", "data"),
     Output("cief-table", "columns")],

    [
        Input("selected-result-ids", "data"),
        Input("max-peaks", "value"),
        Input("prominence-threshold", "value"),
        Input("valley-search-window", "value"),
        Input("valley-drop-ratio", "value"),
        Input("smoothing-window", "value"),
        Input("smoothing-polyorder", "value"),
        Input("y-axis-scaling", "value"),
        Input("subplot-vertical-spacing", "value"),
        Input("subplot-horizontal-spacing", "value"),
        Input("std-cutoff-time", "value"),
        Input("std-end-cutoff-time", "value"),
        Input("std-promincence-threshold", "value"),
        Input("std-valley-search-window", "value"),
        Input("show-std-regression", "value"),
        Input("x-axis-autoscale", "value"),
        Input("x-axis-min", "value"),
        Input("x-axis-max", "value"),
        Input("shade-peaks", "value"),
        Input("pi-method", "value"),
        Input("num-subplot-cols", "value"),

        State("selected-report", "data"),

    ]
)
def reduced_callback(result_ids, max_peaks,
                     prominence_threshold, valley_search_window, valley_drop_ratio,
                     smoothing_window, smoothing_polyorder, y_scale, subplot_vertical_spacing,
                     subplot_horizonatal_spacing, std_cutoff_time,
                     std_end_cutoff_time, std_prominence_threshold, std_valley_search_window, plot_std_regression,
                     x_axis_autoscale, x_axis_min, x_axis_max, shade_peaks, pi_method, num_subplot_cols,
                     selected_report):
    metas = CIEFMetadata.objects.filter(id__in=result_ids)
    result_df_by_id = {
        str(m.id): {
            "sample_id": m.sample_id_full,
            "data": pd.DataFrame(list(
                CIEFTimeSeries.objects.filter(metadata_id=m.id).values("time_min", "channel_1")
            )).sort_values("time_min")
        }
        for m in metas
    }

    # ✅ Initialize table_output as empty list
    table_output = []

    fig, table_data = generate_chromatogram_figure_advanced(
        result_df_by_id,
        # title="Reduced Chromatograms with Peak Classification",
        max_peaks=max_peaks,
        prominence_threshold=prominence_threshold,
        valley_search_window=valley_search_window,
        valley_drop_ratio=valley_drop_ratio,
        smoothing_window=smoothing_window,
        smoothing_polyorder=smoothing_polyorder,
        table_output=table_output,
        y_scale=y_scale,
        subplot_vertical_spacing=subplot_vertical_spacing,
        subplot_horizontal_spacing=subplot_horizonatal_spacing,
        baseline_cutoff_time=std_cutoff_time,
        x_axis_autoscale=x_axis_autoscale,
        x_axis_min=x_axis_min,
        x_axis_max=x_axis_max,
        plot_std_regression=plot_std_regression,
        std_prominence_threshold=std_prominence_threshold,
        std_valley_search_window=std_valley_search_window,
        std_end_cutoff_time=std_end_cutoff_time,
        pi_method=pi_method,
        shade_peaks=shade_peaks,
        num_subplot_cols=num_subplot_cols
    )
    print(f'table output: {table_output}')

    columns = [{"name": k, "id": k} for k in table_data[0].keys()] if table_data else []

    plot_config = {
        'toImageButtonOptions': {
            'filename': f"{datetime.now().strftime('%Y%m%d')}-R-{selected_report}",
            'format': 'png',
            # 'height': 600,
            # 'width': 800,
            # 'scale': 2
        }}
    return fig, plot_config, table_data, columns


@app.callback(
    Output("download-cief-xlsx", "data"),
    Input("export-cief-btn", "n_clicks"),
    State("cief-table", "data"),
    State("selected-report", "data"),
    prevent_initial_call=True
)
def export_table(n_clicks, table_data, report_name):
    filename = f"{report_name}_cIEF.xlsx"
    df = pd.DataFrame(table_data)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1)
        worksheet = writer.sheets["Sheet1"]
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        worksheet.cell(row=1, column=1).value = "cIEF"

    buffer.seek(0)

    def write_buffer(out_io):
        out_io.write(buffer.getvalue())

    return dcc.send_bytes(write_buffer, filename)
