from datetime import datetime

import pandas as pd
import numpy as np
from dash import dcc, html, Input, Output, State, dash_table
from django_plotly_dash import DjangoDash
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from scipy.stats import linregress
from plotly_integration.models import CESDSReport, CESDSMetadata, CESDSTimeSeries
from scipy.signal import find_peaks, savgol_filter, argrelextrema
import dash_bootstrap_components as dbc
from openpyxl import load_workbook
import io

app = DjangoDash("CESDSReportViewerApp")

app.layout = html.Div([
    dcc.Store(id="selected-result-ids"),
    dcc.Store(id="reduced-result-ids"),
    dcc.Store(id="nonreduced-result-ids"),
    dcc.Store(id="standard-regression-params"),
    dcc.Store(id="selected-report"),

    dcc.Tabs(id="main-tabs", value="tab-select-report", persistence=False, children=[

        # Select Report Tab
        dcc.Tab(label="Select Report", value="tab-select-report", children=[
            html.Div([
                dash_table.DataTable(
                    id="cesds-report-table",
                    columns=[
                        {"name": "Report ID", "id": "id"},
                        {"name": "Report Name", "id": "report_name"},
                        {"name": "Project ID", "id": "project_id"},
                        {"name": "User ID", "id": "user_id"},
                        {"name": "Date Created", "id": "date_created"},
                        {"name": "Samples", "id": "num_samples"}
                    ],
                    data=[],
                    row_selectable="single",
                    filter_action="native",
                    sort_action="native",
                    page_size=20,
                    style_table={"height": "70vh", "overflowY": "auto"},
                    style_cell={"textAlign": "center", "padding": "8px"},
                    style_header={
                        "backgroundColor": "#e9f1fb",
                        "fontWeight": "bold",
                        "color": "#0047b3",
                        "borderBottom": "2px solid #0047b3"
                    }
                )
            ])
        ]),

        # Reduced Tab
        dcc.Tab(label="Reduced", value="tab-reduced", children=[
            html.Div([
                dcc.Tabs(id="reduced-subtabs", value="tab-reduced-chrom", persistence=True, children=[
                    dcc.Tab(label="Chromatogram", value="tab-reduced-chrom", children=[
                        html.Div(style={"display": "flex"}, children=[
                            html.Div([
                                dcc.Graph(
                                    id="reduced-chromatogram",
                                    config={'responsive': True},
                                    style={"height": "100%"}
                                )
                            ], style={"width": "80%", "padding": "10px", "minHeight": "1000px"}),

                            html.Div([
                                html.H4("Marker Settings"),
                                html.Label("Marker RT (min):"),
                                dcc.Input(id="marker-rt", type="number", value=7, step=0.1, style={"width": "100%"}),

                                html.Label("Marker Label:"),
                                dcc.Input(id="marker-label", type="text", value="10 kDa",
                                          style={"width": "100%", "marginBottom": "20px"}),

                                html.Hr(),
                                html.H4("Plot Settings"),
                                html.Label("Y- Axis Scaling"),
                                dcc.Input(id="reduced-y-axis-scaling", type="number", value=1.0, step=0.01,
                                          style={"width": "100%"}),
                                html.Label("Subplot Vertical Spacing"),
                                dcc.Input(id="reduced-subplot-vertical-spacing", type="number", value=0.025, step=0.005,
                                          style={"width": "100%"}),

                                html.Hr(),
                                html.H4("Peak Detection"),

                                html.Label("Skip Time After Marker (min):"),
                                dcc.Input(id="skip-time", type="number", value=0.3, step=0.05, style={"width": "100%"}),

                                html.Label("Max Peaks:"),
                                dcc.Input(id="max-peaks", type="number", value=4, step=1, min=1,
                                          style={"width": "100%"}),

                                html.Label("Prominence Threshold:"),
                                dcc.Input(id="prominence-threshold", type="number", value=100.0, step=0.01,
                                          style={"width": "100%"}),

                                html.Label("Valley Search Window (min):"),
                                dcc.Input(id="valley-search-window", type="number", value=1.5, step=0.1,
                                          style={"width": "100%"}),

                                html.Label("Valley Drop Ratio (0–1):"),
                                dcc.Input(id="valley-drop-ratio", type="number", value=0.2, step=0.05, min=0, max=1,
                                          style={"width": "100%"}),

                                html.Label("Smoothing Window (odd integer):"),
                                dcc.Input(id="smoothing-window", type="number", value=3, step=2, min=3,
                                          style={"width": "100%"}),

                                html.Label("Smoothing Polyorder:"),
                                dcc.Input(id="smoothing-polyorder", type="number", value=1, step=1, min=1,
                                          style={"width": "100%", "marginBottom": "20px"}),

                                html.Hr(),
                                html.H4("Light Chain Timing"),
                                html.Button("Calculate Light Chain Time", id="calc-light-chain-btn",
                                            style={"width": "100%", "marginBottom": "10px"}),
                                dcc.Input(id="light-chain-time", type="number", placeholder="Light Chain Time (min)",
                                          readOnly=True, style={"width": "100%"})

                            ], style={"width": "20%", "padding": "10px"})
                        ])
                    ]),
                    dcc.Tab(label="Results Table", value="tab-reduced-table", children=[
                        html.Div([
                            html.H4("Reduced Results Table", style={'textAlign': 'center', 'color': '#0056b3'}),
                            dash_table.DataTable(
                                id="reduced-table",
                                columns=[],  # will be filled by callback
                                data=[],
                                style_header={
                                    'backgroundColor': '#0056b3',  # blue
                                    'color': 'white',
                                    'fontWeight': 'bold',
                                    'textAlign': 'center'
                                },
                                style_table={"overflowX": "auto"},
                                style_cell={"textAlign": "center"},
                            ),
                            html.Button("Export Reduced Table", id="export-reduced-btn"),
                            dcc.Download(id="download-reduced-xlsx")
                        ]),

                    ])
                ])
            ])
        ]),

        # Non-Reduced Tab
        dcc.Tab(label="Non-Reduced", value="tab-nonreduced", children=[
            html.Div([
                dcc.Tabs(id="nonreduced-subtabs", value="tab-nonreduced-chrom", persistence=True, children=[
                    dcc.Tab(label="Chromatogram", value="tab-nonreduced-chrom", children=[
                        html.Div(style={"display": "flex"}, children=[
                            html.Div([
                                dcc.Graph(
                                    id="nonreduced-chromatogram",
                                    config={'responsive': False},
                                    # style={"height": "100%"}
                                )
                            ], style={"width": "80%", "padding": "10px", "minHeight": "1000px"}),

                            html.Div([
                                html.H4("Marker Settings"),
                                html.Label("Marker RT (min):"),
                                dcc.Input(id="nr-marker-rt", type="number", value=7, step=0.1, style={"width": "100%"}),

                                html.Label("Marker Label:"),
                                dcc.Input(id="nr-marker-label", type="text", value="10 kDa",
                                          style={"width": "100%", "marginBottom": "20px"}),

                                html.Hr(),
                                html.H4("Plot Settings"),
                                html.Label("Y- Axis Scaling"),
                                dcc.Input(id="non-reduced-y-axis-scaling", type="number", value=1.0, step=0.01,
                                          style={"width": "100%"}),

                                html.Label("Subplot Vertical Spacing"),
                                dcc.Input(id="non-reduced-subplot-vertical-spacing", type="number", value=0.025,
                                          step=0.005,
                                          style={"width": "100%"}),

                                html.Hr(),
                                html.H4("Peak Detection"),

                                html.Label("Skip Time After Marker (min):"),
                                dcc.Input(id="nr-skip-time", type="number", value=0.3, step=0.05,
                                          style={"width": "100%"}),

                                html.Label("Max Peaks:"),
                                dcc.Input(id="nr-max-peaks", type="number", value=3, step=1, min=1,
                                          style={"width": "100%"}),

                                html.Label("Prominence Threshold:"),
                                dcc.Input(id="nr-prominence-threshold", type="number", value=100, step=0.01,
                                          style={"width": "100%"}),

                                html.Label("Valley Search Window (min):"),
                                dcc.Input(id="nr-valley-search-window", type="number", value=2, step=0.1,
                                          style={"width": "100%"}),

                                html.Label("Valley Drop Ratio (0–1):"),
                                dcc.Input(id="nr-valley-drop-ratio", type="number", value=0.2, step=0.05, min=0, max=1,
                                          style={"width": "100%"}),

                                html.Label("Smoothing Window (odd integer):"),
                                dcc.Input(id="nr-smoothing-window", type="number", value=3, step=2, min=3,
                                          style={"width": "100%"}),

                                html.Label("Smoothing Polyorder:"),
                                dcc.Input(id="nr-smoothing-polyorder", type="number", value=1, step=1, min=1,
                                          style={"width": "100%", "marginBottom": "20px"}),

                                html.Label("Intact Time (optional):"),
                                dcc.Input(id="nr-intact-time", type="number", placeholder="Use tallest peak if blank",
                                          style={"width": "100%", "marginBottom": "20px"}),

                            ], style={"width": "20%", "padding": "10px"})
                        ])
                    ]),
                    dcc.Tab(label="Results Table", value="tab-nonreduced-table", children=[
                        html.Div([
                            html.H4("Non-Reduced Results Table", style={'textAlign': 'center', 'color': '#0056b3'}),
                            dash_table.DataTable(
                                id="nonreduced-table",
                                columns=[],  # will be filled by callback
                                data=[],
                                style_header={
                                    'backgroundColor': '#0056b3',  # blue
                                    'color': 'white',
                                    'fontWeight': 'bold',
                                    'textAlign': 'center'
                                },
                                style_table={"overflowX": "auto"},
                                style_cell={"textAlign": "center"},
                            ),
                            html.Button("Export Reduced Table", id="export-nonreduced-btn"),
                            dcc.Download(id="download-nonreduced-xlsx")
                        ]),

                    ])
                ])
            ])
        ]),
        # Standard Analysis Tab
        dcc.Tab(label="Standard Analysis", value="tab-std-analysis", children=[
            html.Div(
                id='standard-analysis',
                children=[
                    html.H4("Standard Analysis", style={'text-align': 'center', 'color': '#0056b3'}),

                    html.Div([
                        html.Label("Select Standard Sample ID:", style={'color': '#0056b3'}),
                        dcc.Dropdown(
                            id='standard-id-dropdown',
                            placeholder="Select a Standard Sample",
                            style={'width': '100%'}
                        )
                    ], style={'margin-top': '10px'}),

                    dcc.Graph(id='standard-peak-plot'),

                    html.Div([
                        html.P("Regression Equation: ", id="regression-equation"),
                        html.P("R² Value: ", id="r-squared-value"),
                        html.P("Estimated MW for RT: ", id="estimated-mw"),

                        dcc.Input(
                            id="rt-input",
                            type="number",
                            placeholder="Enter Retention Time",
                            style={'width': '80%', 'margin-top': '10px'}
                        ),
                        html.Button("Calculate MW", id="calculate-mw-button", style={
                            'background-color': '#0056b3',
                            'color': 'white',
                            'border': 'none',
                            'padding': '10px',
                            'cursor': 'pointer',
                            'border-radius': '5px'
                        }),
                        dcc.Graph(id='regression-plot', style={'margin-top': '20px'})
                    ], style={
                        'margin-top': '20px',
                        'padding': '10px',
                        'border': '2px solid #0056b3',
                        'border-radius': '5px',
                        'background-color': '#f7f9fc',
                    }),

                    html.Div([
                        html.H4("Detected Peaks & Assigned MWs", style={'color': '#0056b3'}),
                        html.Label("Number of Peaks to Use:"),
                        dcc.Input(
                            id="num-std-peaks",
                            type="number",
                            value=7,
                            min=1,
                            max=20,
                            step=1,
                            style={'width': '100%', 'margin-bottom': '10px'}
                        ),
                        dash_table.DataTable(
                            id="std-detected-peak-table",
                            columns=[
                                {"name": "Retention Time (min)", "id": "peak_rt", "type": "numeric"},
                                {"name": "Peak Height", "id": "peak_height", "type": "numeric"},
                                {"name": "Assigned MW (kDa)", "id": "assigned_mw", "type": "numeric", "editable": True}
                            ],
                            data=[],
                            editable=True,
                            row_selectable="multi",
                            style_table={'overflowX': 'auto'},
                            style_cell={'textAlign': 'center'},
                            style_header={'fontWeight': 'bold', 'backgroundColor': '#e9f1fb'}
                        )
                    ], style={
                        'margin-top': '20px',
                        'padding': '10px',
                        'border': '2px solid #0056b3',
                        'border-radius': '5px',
                        'background-color': '#f7f9fc'
                    })
                ],
                style={'padding': '10px'}
            )
        ])
    ])
])


@app.callback(
    Output("cesds-report-table", "data"),
    Input("main-tabs", "value")
)
def load_report_table(tab_value):
    if tab_value != "tab-select-report":
        return []
    reports = CESDSReport.objects.order_by("-date_created")
    return [{
        "id": r.id,
        "report_name": r.report_name,
        "project_id": r.project_id,
        "user_id": r.user_id,
        "date_created": r.date_created.strftime("%Y-%m-%d %H:%M"),
        "num_samples": len(r.selected_result_ids.split(","))
    } for r in reports]


@app.callback(
    Output("selected-result-ids", "data"),
    Output("selected-report", "data"),
    Input("cesds-report-table", "selected_rows"),
    State("cesds-report-table", "data")
)
def store_selected_result_ids(selected_rows, table_data):
    if selected_rows:
        row = table_data[selected_rows[0]]
        report = CESDSReport.objects.filter(id=row["id"]).first()
        report_name = report.report_name
        if report:
            return [r.strip() for r in report.selected_result_ids.split(",")], report_name
    return [], []


@app.callback(
    [Output("reduced-result-ids", "data"),
     Output("nonreduced-result-ids", "data")],
    Input("selected-result-ids", "data")
)
def split_result_ids_by_prefix(result_ids):
    if not result_ids:
        return [], []

    metas = CESDSMetadata.objects.filter(id__in=result_ids)
    # Filter out samples with 'STD' in the sample name (case insensitive)
    metas = [m for m in metas if not (m.sample_id_full and "std" in m.sample_id_full.lower())]

    reduced = [m.id for m in metas if m.sample_prefix and m.sample_prefix.lower() == "r"]
    nonreduced = [m.id for m in metas if m.sample_prefix and m.sample_prefix.lower() == "nr"]
    return reduced, nonreduced


def detect_valley_to_valley_peaks(
        df,
        signal_col="channel_1",
        time_col="time_min",
        max_peaks=3,
        prominence_threshold=1.0,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3
):
    """
    Detects peaks using valley-to-valley integration with adaptive valley thresholding.
    Returns a list of dictionaries with peak info.
    """
    from scipy.signal import savgol_filter, find_peaks

    time = df[time_col].values
    signal = df[signal_col].values
    smoothed = savgol_filter(signal, window_length=smoothing_window, polyorder=smoothing_polyorder)

    interval = time[1] - time[0]
    min_distance = int(0.3 / interval)
    peak_indices, _ = find_peaks(signal, prominence=prominence_threshold, distance=min_distance)

    peak_indices = sorted(peak_indices, key=lambda i: smoothed[i], reverse=True)[:max_peaks]
    peak_indices.sort()

    peak_infos = []

    for idx in peak_indices:
        peak_time = time[idx]
        peak_height = smoothed[idx]
        min_valley_height = peak_height * (1 - valley_drop_ratio)

        left_limit = max(0, idx - int(valley_search_window / interval))
        left_valley_idx = left_limit + np.argmin(smoothed[left_limit:idx])
        left_valley_val = smoothed[left_valley_idx]

        right_limit = min(len(smoothed) - 1, idx + int(valley_search_window / interval))
        right_valley_idx = idx + np.argmin(smoothed[idx:right_limit + 1])
        right_valley_val = smoothed[right_valley_idx]

        if left_valley_val > min_valley_height or right_valley_val > min_valley_height:
            continue

        baseline = np.linspace(left_valley_val, right_valley_val, right_valley_idx - left_valley_idx + 1)
        signal_segment = smoothed[left_valley_idx:right_valley_idx + 1]
        time_segment = time[left_valley_idx:right_valley_idx + 1]
        area = np.trapz(signal_segment - baseline, time_segment)

        if area > 0:
            peak_infos.append({
                "peak_time": time[idx],
                "peak_height": peak_height,
                "area": area,
                "start_time": time[left_valley_idx],
                "end_time": time[right_valley_idx],
                "baseline": baseline,
                "baseline_time": time_segment,
                "peak_index": idx
            })

    return sorted(peak_infos, key=lambda x: x["area"], reverse=True), smoothed


def shade_peak(fig, df, start_time, end_time, baseline, row, col, label, peak_time, peak_height, class_label):
    PEAK_CLASS_COLORS = {
        "LMW": "rgba(76, 175, 80, 0.3)",  # green (complement of red)
        "Intact": "rgba(33, 150, 243, 0.3)",  # blue (cool, clinical)
        "Light Chain": "rgba(33, 150, 243, 0.3)",  # blue (cool, clinical)
        "Heavy Chain": "rgba(255, 193, 7, 0.3)",  # amber (warm, eye-catching)
        "HMW": "rgba(244, 67, 54, 0.3)",  # red (danger, HMW)

    }

    color = PEAK_CLASS_COLORS.get(class_label, "rgba(0,100,200,0.2)")  # fallback gray

    region = df[(df["time_min"] >= start_time) & (df["time_min"] <= end_time)].copy()
    if region.empty:
        return

    fig.add_trace(
        go.Scatter(
            x=np.concatenate([region["time_min"], region["time_min"][::-1]]),
            y=np.concatenate([region["channel_1"], [baseline] * len(region)]),
            fill="toself",
            fillcolor=color,
            line=dict(color="rgba(255,255,255,0)"),
            showlegend=False,
            hoverinfo="skip"
        ),
        row=row,
        col=col
    )

    fig.add_annotation(
        x=peak_time,
        y=peak_height,
        text=label,
        showarrow=True,
        arrowhead=1,
        ax=0,
        ay=-30,
        row=row,
        col=col
    )

def generate_chromatogram_figure_advanced(
        result_df_by_id,
        title="Chromatograms",
        marker_rt=None,
        marker_label="10 kDa",
        skip_time=0.3,
        max_peaks=4,
        prominence_threshold=0.05,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3,
        light_chain_time=None,
        table_output=None,
        regression_slope=None,
        regression_intercept=None,
        y_scale=1,
        subplot_vertical_spacing=0.25,
):
    if not result_df_by_id:
        return go.Figure(), []

    num_rows = len(result_df_by_id)
    fig = make_subplots(
        rows=num_rows, cols=1,
        shared_xaxes=False,
        vertical_spacing=subplot_vertical_spacing,
        subplot_titles=[meta["sample_id"] for meta in result_df_by_id.values()]
    )

    for i, (mid, meta) in enumerate(result_df_by_id.items(), start=1):
        df = meta["data"]
        if df.empty:
            continue

        fig.add_trace(
            go.Scatter(x=df["time_min"], y=df["channel_1"], mode="lines", name=meta["sample_id"]),
            row=i, col=1
        )

        marker_peak_time = None
        if marker_rt is not None:
            df_marker = df[(df["time_min"] >= marker_rt - 0.5) & (df["time_min"] <= marker_rt + 0.5)]
            if not df_marker.empty:
                idx = df_marker["channel_1"].idxmax()
                marker_peak_time = df_marker.loc[idx, "time_min"]
                peak_height = df_marker.loc[idx, "channel_1"]
                fig.add_annotation(
                    x=marker_peak_time,
                    y=peak_height,
                    text=f"{marker_label} ({marker_peak_time:.2f} min)",
                    showarrow=True,
                    arrowhead=1,
                    ax=0,
                    ay=-40,
                    row=i, col=1
                )

        df_after_marker = df[df["time_min"] > (marker_peak_time or 0) + skip_time].copy()
        peaks, smoothed = detect_valley_to_valley_peaks(
            df_after_marker,
            signal_col="channel_1",
            time_col="time_min",
            max_peaks=max_peaks,
            prominence_threshold=prominence_threshold,
            valley_search_window=valley_search_window,
            valley_drop_ratio=valley_drop_ratio,
            smoothing_window=smoothing_window,
            smoothing_polyorder=smoothing_polyorder,
        )

        peaks = sorted(peaks, key=lambda x: x["peak_height"], reverse=True)
        light_chain = heavy_chain = None
        if len(peaks) >= 2:
            top2 = peaks[:2]
            if top2[0]["peak_time"] < top2[1]["peak_time"]:
                light_chain, heavy_chain = top2[0], top2[1]
            else:
                light_chain, heavy_chain = top2[1], top2[0]

        classified_peaks = []

        for p in peaks:
            class_label = None
            if p == light_chain:
                class_label = "Light Chain"
            elif p == heavy_chain:
                class_label = "Heavy Chain"
            elif light_chain and p["peak_time"] < light_chain["peak_time"]:
                class_label = "LMW"
            elif heavy_chain and p["peak_time"] > heavy_chain["peak_time"]:
                class_label = "HMW"

            if class_label:
                classified_peaks.append((p, class_label))

        total_area = sum(p["area"] for p, _ in classified_peaks)
        percentages = {"LMW": 0.0, "Light Chain": 0.0, "Heavy Chain": 0.0, "HMW": 0.0}

        max_signal = df["channel_1"].max()
        if not y_scale or y_scale == 1:
            fig.update_yaxes(title_text="UV", row=i, col=1)
        else:
            y_max = max_signal * y_scale
            fig.update_yaxes(
                title_text="UV",
                autorange=True,
                autorangeoptions=dict(maxallowed=y_max),
                row=i, col=1
            )

        fig.update_xaxes(title_text="Time (min)", row=i, col=1)

        for p, class_label in classified_peaks:
            pct = (p["area"] / total_area * 100) if total_area else 0
            percentages[class_label] += pct
            label = f"{class_label}<br>({pct:.1f}%)"

            if regression_slope is not None and regression_intercept is not None:
                log_mw = regression_slope * p["peak_time"] + regression_intercept
                mw_kda = np.exp(log_mw)
                label += f"<br>{mw_kda:.1f} kDa"

            peak_height = p["peak_height"]
            if y_scale and y_scale != 1:
                peak_height = min(peak_height, y_max)

            shade_peak(
                fig, df,
                start_time=p["start_time"],
                end_time=p["end_time"],
                baseline=p["baseline"][0],
                row=i, col=1,
                label=label,
                peak_time=p["peak_time"],
                peak_height=peak_height,
                class_label=class_label
            )

        total_pct = sum(percentages.values())
        if abs(total_pct - 100) > 1e-2:
            print(f"[DEBUG] {meta['sample_id']} percentages do not sum to 100%: {total_pct:.2f}%")

        if table_output is not None:
            table_output.append({
                "Sample Name": meta["sample_id"],
                "LMW (%)": round(percentages["LMW"], 1),
                "Light Chain (%)": round(percentages["Light Chain"], 1),
                "Heavy Chain (%)": round(percentages["Heavy Chain"], 1),
                "HMW (%)": round(percentages["HMW"], 1),
            })

    fig.update_layout(
        height=300 * num_rows,
        title=title,
        showlegend=False,
        template="plotly_white",
        margin=dict(t=40, b=40, l=40, r=30)
    )
    return fig, table_output


# def generate_chromatogram_figure_advanced(
#         result_df_by_id,
#         title="Chromatograms",
#         marker_rt=None,
#         marker_label="10 kDa",
#         skip_time=0.3,
#         max_peaks=4,
#         prominence_threshold=0.05,
#         valley_search_window=3.0,
#         valley_drop_ratio=0.2,
#         smoothing_window=11,
#         smoothing_polyorder=3,
#         light_chain_time=None,
#         table_output=None,
#         regression_slope=None,
#         regression_intercept=None,
#         y_scale=1,
#         subplot_vertical_spacing=0.25,
# ):
#     num_rows = len(result_df_by_id)
#
#     fig = make_subplots(
#         rows=num_rows, cols=1,
#         shared_xaxes=False,
#         vertical_spacing=subplot_vertical_spacing,
#         subplot_titles=[meta["sample_id"] for meta in result_df_by_id.values()]
#     )
#
#     for i, (mid, meta) in enumerate(result_df_by_id.items(), start=1):
#         df = meta["data"]
#         if df.empty:
#             continue
#
#         fig.add_trace(
#             go.Scatter(x=df["time_min"], y=df["channel_1"], mode="lines", name=meta["sample_id"]),
#             row=i, col=1
#         )
#
#         # Marker annotation
#         marker_peak_time = None
#         if marker_rt is not None:
#             df_marker = df[(df["time_min"] >= marker_rt - 0.5) & (df["time_min"] <= marker_rt + 0.5)]
#             if not df_marker.empty:
#                 idx = df_marker["channel_1"].idxmax()
#                 marker_peak_time = df_marker.loc[idx, "time_min"]
#                 peak_height = df_marker.loc[idx, "channel_1"]
#
#                 fig.add_annotation(
#                     x=marker_peak_time,
#                     y=peak_height,
#                     text=f"{marker_label} ({marker_peak_time:.2f} min)",
#                     showarrow=True,
#                     arrowhead=1,
#                     ax=0,
#                     ay=-40,
#                     row=i, col=1
#                 )
#
#         # Peak detection
#         df_after_marker = df[df["time_min"] > (marker_peak_time or 0) + skip_time].copy()
#         peaks, smoothed = detect_valley_to_valley_peaks(
#             df_after_marker,
#             signal_col="channel_1",
#             time_col="time_min",
#             max_peaks=max_peaks,
#             prominence_threshold=prominence_threshold,
#             valley_search_window=valley_search_window,
#             valley_drop_ratio=valley_drop_ratio,
#             smoothing_window=smoothing_window,
#             smoothing_polyorder=smoothing_polyorder,
#         )
#
#         total_area = sum(p["area"] for p in peaks)
#         peaks = sorted(peaks, key=lambda x: x["peak_height"], reverse=True)
#
#         light_chain = heavy_chain = None
#         if len(peaks) >= 2:
#             top2 = peaks[:2]
#             if top2[0]["peak_time"] < top2[1]["peak_time"]:
#                 light_chain, heavy_chain = top2[0], top2[1]
#             else:
#                 light_chain, heavy_chain = top2[1], top2[0]
#
#         percentages = {"LMW": 0.0, "Light Chain": 0.0, "Heavy Chain": 0.0, "HMW": 0.0}
#
#         for p in peaks:
#
#             # Y Axes Scaling
#             if not y_scale or y_scale == 1:
#                 autoscale = True
#                 fig.update_yaxes(title_text="UV", row=i, col=1)
#             else:
#                 autoscale = False
#                 max_signal = df["channel_1"].max()
#                 y_max = max_signal * y_scale
#                 fig.update_yaxes(
#                     title_text="UV",
#                     autorange=True,
#                     autorangeoptions=dict(maxallowed=y_max),
#                     row=i,
#                     col=1
#                 )
#             fig.update_xaxes(title_text="Time (min)", row=i, col=1)
#
#             # Peak Classification and percentage calculation
#
#             pct = (p["area"] / total_area * 100) if total_area else 0
#             label = f"{pct:.1f}%"
#             class_label = None
#
#             if p == light_chain:
#                 label = f"Light Chain<br>({pct:.1f}%)"
#                 class_label = "Light Chain"
#             elif p == heavy_chain:
#                 label = f"Heavy Chain<br>({pct:.1f}%)"
#                 class_label = "Heavy Chain"
#             elif light_chain and p["peak_time"] < light_chain["peak_time"]:
#                 label = f"LMW<br>({pct:.1f}%)"
#                 class_label = "LMW"
#             elif heavy_chain and p["peak_time"] > heavy_chain["peak_time"]:
#                 label = f"HMW<br>({pct:.1f}%)"
#                 class_label = "HMW"
#             else:
#                 continue
#
#             # Optional: Calculate MW
#             if regression_slope is not None and regression_intercept is not None:
#                 log_mw = regression_slope * p["peak_time"] + regression_intercept
#                 mw_kda = np.exp(log_mw)
#                 label += f"<br>{mw_kda:.1f} kDa"
#
#             percentages[class_label] += pct
#
#             if autoscale:
#                 peak_height = p["peak_height"]
#             elif autoscale == False and p["peak_height"] < y_max:
#                 peak_height = p["peak_height"]
#             else:
#                 peak_height = y_max
#
#             shade_peak(
#                 fig, df,
#                 start_time=p["start_time"],
#                 end_time=p["end_time"],
#                 baseline=p["baseline"][0],
#                 row=i, col=1,
#                 label=label,
#                 peak_time=p["peak_time"],
#                 peak_height=peak_height,
#                 class_label=class_label
#             )
#
#         # Append summary to table output
#         if table_output is not None:
#             table_output.append({
#                 "Sample Name": meta["sample_id"],
#                 "LMW (%)": round(percentages["LMW"], 1),
#                 "Light Chain (%)": round(percentages["Light Chain"], 1),
#                 "Heavy Chain (%)": round(percentages["Heavy Chain"], 1),
#                 "HMW (%)": round(percentages["HMW"], 1),
#             })
#
#     fig.update_layout(
#         height=300 * num_rows,
#         title=title,
#         showlegend=False,
#         template="plotly_white",
#         margin=dict(t=40, b=40, l=40, r=30)
#     )
#     print(f'table output: {table_output}')
#     return fig, table_output


@app.callback(
    [Output("reduced-chromatogram", "figure"),
     Output("reduced-chromatogram", "config"),
     Output("reduced-table", "data"),
     Output("reduced-table", "columns")],

    [
        Input("reduced-result-ids", "data"),
        Input("marker-rt", "value"),
        Input("marker-label", "value"),
        Input("skip-time", "value"),
        Input("max-peaks", "value"),
        Input("prominence-threshold", "value"),
        Input("valley-search-window", "value"),
        Input("valley-drop-ratio", "value"),
        Input("smoothing-window", "value"),
        Input("smoothing-polyorder", "value"),
        Input("light-chain-time", "value"),
        Input("standard-regression-params", "data"),
        Input("reduced-y-axis-scaling", "value"),
        Input("reduced-subplot-vertical-spacing", "value"),
        State("selected-report", "data"),

    ]
)
def reduced_callback(result_ids, marker_rt, marker_label, skip_time, max_peaks,
                     prominence_threshold, valley_search_window, valley_drop_ratio,
                     smoothing_window, smoothing_polyorder, light_chain_time,
                     regression_params, y_scale, subplot_vertical_spacing, selected_report):
    metas = CESDSMetadata.objects.filter(id__in=result_ids)
    result_df_by_id = {}

    for m in metas:
        qs = CESDSTimeSeries.objects.filter(metadata_id=m.id).values("time_min", "channel_1")
        df = pd.DataFrame(list(qs))

        if df.empty or "time_min" not in df.columns or "channel_1" not in df.columns:
            continue

        result_df_by_id[str(m.id)] = {
            "sample_id": m.sample_id_full,
            "data": df.sort_values("time_min")
        }

    # ✅ Handle the empty case after building the dict
    if not result_df_by_id:
        return go.Figure(), {}, [], []

    # ✅ Initialize table_output as empty list
    table_output = []

    slope = regression_params.get("slope") if regression_params else None
    intercept = regression_params.get("intercept") if regression_params else None

    fig, table_data = generate_chromatogram_figure_advanced(
        result_df_by_id,
        title="Reduced Chromatograms with Peak Classification",
        marker_rt=marker_rt,
        marker_label=marker_label,
        skip_time=skip_time,
        max_peaks=max_peaks,
        prominence_threshold=prominence_threshold,
        valley_search_window=valley_search_window,
        valley_drop_ratio=valley_drop_ratio,
        smoothing_window=smoothing_window,
        smoothing_polyorder=smoothing_polyorder,
        light_chain_time=light_chain_time,
        table_output=table_output,
        regression_slope=slope,
        regression_intercept=intercept,
        y_scale=y_scale,
        subplot_vertical_spacing=subplot_vertical_spacing
    )
    print(f'table output: {table_output}')

    columns = [{"name": k, "id": k} for k in table_data[0].keys()] if table_data else []

    plot_config = {
        'toImageButtonOptions': {
            'filename': f"{datetime.now().strftime('%Y%m%d')}-R-{selected_report}",
            'format': 'png',
            # 'height': 600,
            # 'width': 800,
            # 'scale': 2
        }}
    return fig, plot_config, table_data, columns


@app.callback(
    Output("download-reduced-xlsx", "data"),
    Input("export-reduced-btn", "n_clicks"),
    State("reduced-table", "data"),
    State("selected-report", "data"),
    prevent_initial_call=True
)
def export_nonreduced_table(n_clicks, table_data, report_name):
    filename = f"{report_name}_Reduced.xlsx"
    df = pd.DataFrame(table_data)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1)
        worksheet = writer.sheets["Sheet1"]
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        worksheet.cell(row=1, column=1).value = "Reduced Results"

    buffer.seek(0)

    def write_buffer(out_io):
        out_io.write(buffer.getvalue())

    return dcc.send_bytes(write_buffer, filename)


# def generate_chromatogram_figure_nonreduced(
#         result_df_by_id,
#         title="Non-Reduced Chromatograms",
#         marker_rt=None,
#         marker_label="10 kDa",
#         skip_time=0.3,
#         max_peaks=4,
#         prominence_threshold=0.05,
#         valley_search_window=3.0,
#         valley_drop_ratio=0.2,
#         smoothing_window=11,
#         smoothing_polyorder=3,
#         intact_time=None,
#         table_output=None,
#         regression_slope=None,
#         regression_intercept=None,
#         y_scale=1,
#         subplot_vertical_spacing=0.25,
# ):
#     num_rows = len(result_df_by_id)
#     fig = make_subplots(
#         rows=num_rows, cols=1,
#         shared_xaxes=False,
#         vertical_spacing=subplot_vertical_spacing,
#         subplot_titles=[meta["sample_id"] for meta in result_df_by_id.values()]
#     )
#
#     for i, (mid, meta) in enumerate(result_df_by_id.items(), start=1):
#         df = meta["data"]
#         if df.empty:
#             continue
#
#         fig.add_trace(
#             go.Scatter(x=df["time_min"], y=df["channel_1"], mode="lines", name=meta["sample_id"]),
#             row=i, col=1
#         )
#
#         marker_peak_time = None
#         if marker_rt is not None:
#             df_marker = df[(df["time_min"] >= marker_rt - 0.5) & (df["time_min"] <= marker_rt + 0.5)]
#             if not df_marker.empty:
#                 idx = df_marker["channel_1"].idxmax()
#                 marker_peak_time = df_marker.loc[idx, "time_min"]
#                 peak_height = df_marker.loc[idx, "channel_1"]
#
#                 fig.add_annotation(
#                     x=marker_peak_time,
#                     y=peak_height,
#                     text=f"{marker_label} ({marker_peak_time:.2f} min)",
#                     showarrow=True,
#                     arrowhead=1,
#                     ax=0,
#                     ay=-40,
#                     row=i, col=1
#                 )
#
#         df_after_marker = df[df["time_min"] > (marker_peak_time or 0) + skip_time].copy()
#         peaks, _ = detect_valley_to_valley_peaks(
#             df_after_marker,
#             signal_col="channel_1",
#             time_col="time_min",
#             max_peaks=max_peaks,
#             prominence_threshold=prominence_threshold,
#             valley_search_window=valley_search_window,
#             valley_drop_ratio=valley_drop_ratio,
#             smoothing_window=smoothing_window,
#             smoothing_polyorder=smoothing_polyorder,
#         )
#
#         peaks = sorted(peaks, key=lambda x: x["peak_time"])
#         total_area = sum(p["area"] for p in peaks)
#
#         # Determine intact peak
#         intact = None
#         if intact_time:
#             closest = min(peaks, key=lambda p: abs(p["peak_time"] - intact_time))
#             intact = closest
#         elif peaks:
#             intact = max(peaks, key=lambda p: p["peak_height"])
#
#         percentages = {"LMW": 0.0, "Light Chain": 0.0, "Intact": 0.0, "HMW": 0.0}
#
#         before_intact = [p for p in peaks if p["peak_time"] < intact["peak_time"]] if intact else []
#         after_intact = [p for p in peaks if p["peak_time"] > intact["peak_time"]] if intact else []
#
#         for p in peaks:
#             # Y Axes Scaling
#             if not y_scale or y_scale == 1:
#                 autoscale = True
#                 fig.update_yaxes(title_text="UV", row=i, col=1)
#             else:
#                 autoscale = False
#                 max_signal = df["channel_1"].max()
#                 y_max = max_signal * y_scale
#                 fig.update_yaxes(
#                     title_text="UV",
#                     autorange=True,
#                     autorangeoptions=dict(maxallowed=y_max),
#                     row=i,
#                     col=1
#                 )
#             fig.update_xaxes(title_text="Time (min)", row=i, col=1)
#
#             pct = (p["area"] / total_area * 100) if total_area else 0
#             class_label = None
#
#             if intact and p["peak_time"] == intact["peak_time"]:
#                 class_label = "Intact"
#             elif p in after_intact:
#                 class_label = "HMW"
#             elif p in before_intact:
#                 if len(before_intact) == 1:
#                     class_label = "LMW"
#                 elif len(before_intact) >= 2:
#                     class_label = "LMW" if p == before_intact[0] else "Light Chain"
#
#             if not class_label:
#                 continue
#
#             label = f"{class_label}<br>({pct:.1f}%)"
#
#             # 🧠 Add MW annotation
#             if regression_slope is not None and regression_intercept is not None:
#                 log_mw = regression_slope * p["peak_time"] + regression_intercept
#                 mw_kda = np.exp(log_mw)
#                 label += f"<br>{mw_kda:.1f} kDa"
#
#             percentages[class_label] += pct
#
#             if autoscale:
#                 peak_height = p["peak_height"]
#             elif autoscale == False and p["peak_height"] < y_max:
#                 peak_height = p["peak_height"]
#             else:
#                 peak_height = y_max
#
#             shade_peak(
#                 fig, df,
#                 start_time=p["start_time"],
#                 end_time=p["end_time"],
#                 baseline=p["baseline"][0],
#                 row=i, col=1,
#                 label=label,
#                 peak_time=p["peak_time"],
#                 peak_height=peak_height,
#                 class_label=class_label
#             )
#
#         if table_output is not None:
#             table_output.append({
#                 "Sample Name": meta["sample_id"],
#                 "LMW (%)": round(percentages["LMW"], 1),
#                 "Light Chain (%)": round(percentages["Light Chain"], 1),
#                 "Intact (%)": round(percentages["Intact"], 1),
#                 "HMW (%)": round(percentages["HMW"], 1),
#             })
#
#     fig.update_layout(
#         height=300 * num_rows,
#         title=title,
#         showlegend=False,
#         template="plotly_white",
#         margin=dict(t=40, b=40, l=40, r=30)
#     )
#
#     return fig, table_output

def generate_chromatogram_figure_nonreduced(
        result_df_by_id,
        title="Non-Reduced Chromatograms",
        marker_rt=None,
        marker_label="10 kDa",
        skip_time=0.3,
        max_peaks=4,
        prominence_threshold=0.05,
        valley_search_window=3.0,
        valley_drop_ratio=0.2,
        smoothing_window=11,
        smoothing_polyorder=3,
        intact_time=None,
        table_output=None,
        regression_slope=None,
        regression_intercept=None,
        y_scale=1,
        subplot_vertical_spacing=0.25,
):
    if not result_df_by_id:
        return go.Figure(), []

    num_rows = len(result_df_by_id)
    fig = make_subplots(
        rows=num_rows, cols=1,
        shared_xaxes=False,
        vertical_spacing=subplot_vertical_spacing,
        subplot_titles=[meta["sample_id"] for meta in result_df_by_id.values()]
    )

    for i, (mid, meta) in enumerate(result_df_by_id.items(), start=1):
        df = meta["data"]
        if df.empty:
            continue

        fig.add_trace(
            go.Scatter(x=df["time_min"], y=df["channel_1"], mode="lines", name=meta["sample_id"]),
            row=i, col=1
        )

        marker_peak_time = None
        if marker_rt is not None:
            df_marker = df[(df["time_min"] >= marker_rt - 0.5) & (df["time_min"] <= marker_rt + 0.5)]
            if not df_marker.empty:
                idx = df_marker["channel_1"].idxmax()
                marker_peak_time = df_marker.loc[idx, "time_min"]
                peak_height = df_marker.loc[idx, "channel_1"]

                fig.add_annotation(
                    x=marker_peak_time,
                    y=peak_height,
                    text=f"{marker_label} ({marker_peak_time:.2f} min)",
                    showarrow=True,
                    arrowhead=1,
                    ax=0,
                    ay=-40,
                    row=i, col=1
                )

        df_after_marker = df[df["time_min"] > (marker_peak_time or 0) + skip_time].copy()
        peaks, _ = detect_valley_to_valley_peaks(
            df_after_marker,
            signal_col="channel_1",
            time_col="time_min",
            max_peaks=max_peaks,
            prominence_threshold=prominence_threshold,
            valley_search_window=valley_search_window,
            valley_drop_ratio=valley_drop_ratio,
            smoothing_window=smoothing_window,
            smoothing_polyorder=smoothing_polyorder,
        )

        peaks = sorted(peaks, key=lambda x: x["peak_time"])
        intact = None
        if intact_time:
            closest = min(peaks, key=lambda p: abs(p["peak_time"] - intact_time))
            intact = closest
        elif peaks:
            intact = max(peaks, key=lambda p: p["peak_height"])

        classified_peaks = []
        before_intact = [p for p in peaks if p["peak_time"] < intact["peak_time"]] if intact else []
        after_intact = [p for p in peaks if p["peak_time"] > intact["peak_time"]] if intact else []

        for p in peaks:
            class_label = None
            if intact and p["peak_time"] == intact["peak_time"]:
                class_label = "Intact"
            elif p in after_intact:
                class_label = "HMW"
            elif p in before_intact:
                if len(before_intact) == 1:
                    class_label = "LMW"
                elif len(before_intact) >= 2:
                    class_label = "LMW" if p == before_intact[0] else "Light Chain"

            if class_label:
                classified_peaks.append((p, class_label))

        total_area = sum(p["area"] for p, _ in classified_peaks)
        percentages = {"LMW": 0.0, "Light Chain": 0.0, "Intact": 0.0, "HMW": 0.0}

        max_signal = df["channel_1"].max()
        if not y_scale or y_scale == 1:
            fig.update_yaxes(title_text="UV", row=i, col=1)
        else:
            y_max = max_signal * y_scale
            fig.update_yaxes(
                title_text="UV",
                autorange=True,
                autorangeoptions=dict(maxallowed=y_max),
                row=i, col=1
            )

        fig.update_xaxes(title_text="Time (min)", row=i, col=1)

        for p, class_label in classified_peaks:
            pct = (p["area"] / total_area * 100) if total_area else 0
            percentages[class_label] += pct
            label = f"{class_label}<br>({pct:.1f}%)"

            if regression_slope is not None and regression_intercept is not None:
                log_mw = regression_slope * p["peak_time"] + regression_intercept
                mw_kda = np.exp(log_mw)
                label += f"<br>{mw_kda:.1f} kDa"

            peak_height = p["peak_height"]
            if y_scale and y_scale != 1:
                peak_height = min(peak_height, y_max)

            shade_peak(
                fig, df,
                start_time=p["start_time"],
                end_time=p["end_time"],
                baseline=p["baseline"][0],
                row=i, col=1,
                label=label,
                peak_time=p["peak_time"],
                peak_height=peak_height,
                class_label=class_label
            )

        total_pct = sum(percentages.values())
        if abs(total_pct - 100) > 1e-2:
            print(f"[DEBUG] {meta['sample_id']} percentages do not sum to 100%: {total_pct:.2f}%")

        if table_output is not None:
            table_output.append({
                "Sample Name": meta["sample_id"],
                "LMW (%)": round(percentages["LMW"], 1),
                "Light Chain (%)": round(percentages["Light Chain"], 1),
                "Intact (%)": round(percentages["Intact"], 1),
                "HMW (%)": round(percentages["HMW"], 1),
            })

    fig.update_layout(
        height=300 * num_rows,
        title=title,
        showlegend=False,
        template="plotly_white",
        margin=dict(t=40, b=40, l=40, r=30)
    )

    return fig, table_output



@app.callback(
    [
        Output("nonreduced-chromatogram", "figure"),
        Output("nonreduced-chromatogram", "config"),
        Output("nonreduced-table", "data"),
        Output("nonreduced-table", "columns")
    ],
    [
        Input("nonreduced-result-ids", "data"),
        Input("nr-marker-rt", "value"),
        Input("nr-marker-label", "value"),
        Input("nr-skip-time", "value"),
        Input("nr-max-peaks", "value"),
        Input("nr-prominence-threshold", "value"),
        Input("nr-valley-search-window", "value"),
        Input("nr-valley-drop-ratio", "value"),
        Input("nr-smoothing-window", "value"),
        Input("nr-smoothing-polyorder", "value"),
        Input("nr-intact-time", "value"),
        Input("standard-regression-params", "data"),
        Input("non-reduced-y-axis-scaling", "value"),
        Input("non-reduced-subplot-vertical-spacing", "value"),
        State("selected-report", "data"),
    ]
)
def nonreduced_callback(result_ids, marker_rt, marker_label, skip_time, max_peaks,
                        prominence_threshold, valley_search_window, valley_drop_ratio,
                        smoothing_window, smoothing_polyorder, intact_time, regression_params, y_scale,
                        subplot_vertical_spacing, selected_report):


    metas = CESDSMetadata.objects.filter(id__in=result_ids)
    result_df_by_id = {}

    for m in metas:
        qs = CESDSTimeSeries.objects.filter(metadata_id=m.id).values("time_min", "channel_1")
        df = pd.DataFrame(list(qs))

        if df.empty or "time_min" not in df.columns or "channel_1" not in df.columns:
            continue

        result_df_by_id[str(m.id)] = {
            "sample_id": m.sample_id_full,
            "data": df.sort_values("time_min")
        }

    # ✅ Handle the empty case after building the dict
    if not result_df_by_id:
        return go.Figure(), {}, [], []

    table_output = []
    slope = regression_params.get("slope") if regression_params else None
    intercept = regression_params.get("intercept") if regression_params else None

    fig, table_data = generate_chromatogram_figure_nonreduced(
        result_df_by_id,
        title="Non-Reduced Chromatograms with Peak Classification",
        marker_rt=marker_rt,
        marker_label=marker_label,
        skip_time=skip_time,
        max_peaks=max_peaks,
        prominence_threshold=prominence_threshold,
        valley_search_window=valley_search_window,
        valley_drop_ratio=valley_drop_ratio,
        smoothing_window=smoothing_window,
        smoothing_polyorder=smoothing_polyorder,
        intact_time=intact_time,
        table_output=table_output,
        regression_slope=slope,
        regression_intercept=intercept,
        y_scale=y_scale,
        subplot_vertical_spacing=subplot_vertical_spacing
    )

    columns = [{"name": k, "id": k} for k in table_data[0].keys()] if table_data else []

    print(f'table output: {table_output}')

    plot_config = {
        'toImageButtonOptions': {
            'filename': f"{datetime.now().strftime('%Y%m%d')}-NR-{selected_report}",
            'format': 'png',
            # 'height': 600,
            # 'width': 800,
            # 'scale': 2
        }}

    return fig, plot_config, table_data, columns


@app.callback(
    Output("download-nonreduced-xlsx", "data"),
    Input("export-nonreduced-btn", "n_clicks"),
    State("nonreduced-table", "data"),
    State("selected-report", "data"),
    prevent_initial_call=True
)
def export_nonreduced_table(n_clicks, table_data, report_name):
    filename = f"{report_name}_NonReduced.xlsx"
    df = pd.DataFrame(table_data)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1)
        worksheet = writer.sheets["Sheet1"]
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        worksheet.cell(row=1, column=1).value = "Non-Reduced Results"

    buffer.seek(0)

    def write_buffer(out_io):
        out_io.write(buffer.getvalue())

    return dcc.send_bytes(write_buffer, filename)


# Standard Analysis Logic
@app.callback(
    Output("standard-id-dropdown", "options"),
    Input("selected-result-ids", "data")
)
def get_std_dropdown_options(result_ids):
    if not result_ids:
        return []

    stds = CESDSMetadata.objects.filter(
        id__in=result_ids,
        sample_id_full__iregex=r".*STD.*"
    )

    return [{"label": s.sample_id_full, "value": s.id} for s in stds]


@app.callback(
    Output("standard-id-dropdown", "value"),
    Input("standard-id-dropdown", "options")
)
def auto_select_first_std(options):
    if options:
        return options[0]["value"]
    return None


@app.callback(
    [
        Output("standard-peak-plot", "figure"),
        Output("std-detected-peak-table", "data")
    ],
    [
        Input("standard-id-dropdown", "value"),
        State("num-std-peaks", "value")
    ]
)
def plot_std_chromatogram_and_generate_table(metadata_id, num_to_keep):
    if not metadata_id:
        return go.Figure(), []

    qs = CESDSTimeSeries.objects.filter(metadata_id=metadata_id).values("time_min", "channel_1")
    df = pd.DataFrame(list(qs)).sort_values("time_min")
    if df.empty:
        return go.Figure(), []

    # Plot full chromatogram
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["time_min"],
        y=df["channel_1"],
        mode="lines",
        name="STD Chromatogram"
    ))

    # Filter to >5 min for peak detection
    df_filtered = df[df["time_min"] > 5.0]
    if df_filtered.empty:
        return fig, []

    x = df_filtered["time_min"].values
    y = df_filtered["channel_1"].values
    peak_indices, _ = find_peaks(y, distance=5)
    all_peaks = [{"peak_rt": x[i], "peak_height": y[i]} for i in peak_indices]

    # Step 1: Take top N peaks by height
    top_peaks = sorted(all_peaks, key=lambda p: p["peak_height"], reverse=True)[:num_to_keep or 7]

    # Step 2: Sort selected peaks by RT (ascending)
    sorted_peaks = sorted(top_peaks, key=lambda p: p["peak_rt"])

    # Step 3: Assign MWs linearly (left = 10 kDa, right = 250 kDa)
    default_mws = [10, 20, 35, 50, 100, 150, 250]
    assigned_mws = default_mws[:len(sorted_peaks)]

    table_data = []
    for i, peak in enumerate(sorted_peaks):
        table_data.append({
            "peak_rt": round(peak["peak_rt"], 3),
            "peak_height": round(peak["peak_height"], 3),
            "assigned_mw": assigned_mws[i] if i < len(assigned_mws) else ""
        })

    # Add peak markers to plot
    fig.add_trace(go.Scatter(
        x=[p["peak_rt"] for p in sorted_peaks],
        y=[p["peak_height"] for p in sorted_peaks],
        mode="markers+text",
        text=[f"{i + 1}" for i in range(len(sorted_peaks))],
        textposition="top center",
        marker=dict(color="red", size=8),
        name="Top Peaks"
    ))

    fig.update_layout(
        title="Standard Chromatogram with Assigned Peaks",
        xaxis_title="Time (min)",
        yaxis_title="UV (Channel 1)",
        template="plotly_white"
    )

    return fig, table_data


@app.callback(
    [
        Output("regression-equation", "children"),
        Output("r-squared-value", "children"),
        Output("regression-plot", "figure"),
        Output("estimated-mw", "children"),
        Output("standard-regression-params", "data")
    ],
    [
        Input("std-detected-peak-table", "data"),
        Input("std-detected-peak-table", "selected_rows"),
        Input("rt-input", "value")
    ]
)
def run_linear_mw_regression(table_data, selected_rows, rt_input):
    if not table_data or not selected_rows:
        return "No selected points", "N/A", go.Figure(), "N/A", {}

    df = pd.DataFrame(table_data)
    df = df.iloc[selected_rows]  # Only keep selected rows
    df = df.dropna(subset=["peak_rt", "assigned_mw"])
    df = df[df["assigned_mw"] > 0]

    if df.shape[0] < 2:
        return "Select ≥2 points", "N/A", go.Figure(), "N/A", {}

    try:
        x = df["peak_rt"]
        y = np.log(df["assigned_mw"])
        slope, intercept, r_value, _, _ = linregress(x, y)

        # Generate regression line
        x_vals = np.linspace(x.min(), x.max(), 100)
        y_vals = slope * x_vals + intercept

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=x,
            y=np.log(df["assigned_mw"]),  # y = log(MW)
            mode="markers+text",
            text=[f"{mw:.0f} kDa" for mw in df["assigned_mw"]],  # ✅ correct label from original MWs
            textposition="top center",
            name="Selected Points"
        ))
        fig.add_trace(go.Scatter(
            x=x_vals,
            y=y_vals,
            mode="lines",
            name="Regression Line"
        ))
        fig.update_layout(
            title="Log MW vs Retention Time",
            xaxis_title="Retention Time (min)",
            yaxis_title="Log MW (kDa)",
            template="plotly_white"
        )

        estimated_mw = "N/A"
        if rt_input is not None:
            log_mw = slope * rt_input + intercept
            estimated_mw = f"{np.exp(log_mw) / 1000:.2f} kD"

        return (
            f"MW = {slope:.3f} × RT + {intercept:.3f}",
            f"R² = {r_value ** 2:.4f}",
            fig,
            estimated_mw,
            {"slope": slope, "intercept": intercept}
        )

    except Exception as e:
        return f"Error: {e}", "N/A", go.Figure(), "N/A", {}


@app.callback(
    Output("std-detected-peak-table", "selected_rows"),
    Input("std-detected-peak-table", "data")
)
def auto_select_all_std_peaks(data):
    return list(range(len(data))) if data else []
